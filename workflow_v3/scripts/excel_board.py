#!/usr/bin/env python3
import contextlib
import os
import pathlib
import tempfile
from typing import Callable, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SHEET_NAME = "workflow_v3"

COL_MK_PATH = "原始mk/makefile路径"
COL_CMAKE_PATH = "转换后cmake文件路径"
COL_ANALYSIS_DONE = "是否完成分析"
COL_ANALYSIS_OK = "分析是否成功"
COL_CONVERSION_DONE = "是否完成转换"
COL_CONVERSION_OK = "转换是否成功"
COL_MANUAL_COMMENT = "人工意见"

COL_TASK_ID = "任务ID"
COL_ANALYSIS_STATUS = "分析状态"
COL_ANALYSIS_RETURN = "分析返回码"
COL_ANALYSIS_TARGETS = "分析目标数量"
COL_ANALYSIS_RESULT = "分析结果json路径"
COL_ANALYSIS_STDOUT = "分析stdout路径"
COL_ANALYSIS_STDERR = "分析stderr路径"
COL_ANALYSIS_UPDATED = "分析更新时间"
COL_CONVERSION_STATUS = "转换状态"
COL_CONVERSION_RETURN = "转换返回码"
COL_CONVERSION_STDOUT = "转换stdout路径"
COL_CONVERSION_STDERR = "转换stderr路径"
COL_CONVERSION_UPDATED = "转换更新时间"
COL_RETRY_COUNT = "转换重试次数"

REQUIRED_COLUMNS = [
    COL_MK_PATH,
    COL_CMAKE_PATH,
    COL_ANALYSIS_DONE,
    COL_ANALYSIS_OK,
    COL_CONVERSION_DONE,
    COL_CONVERSION_OK,
    COL_MANUAL_COMMENT,
]

DEFAULT_COLUMNS = REQUIRED_COLUMNS + [
    COL_TASK_ID,
    COL_ANALYSIS_STATUS,
    COL_ANALYSIS_RETURN,
    COL_ANALYSIS_TARGETS,
    COL_ANALYSIS_RESULT,
    COL_ANALYSIS_STDOUT,
    COL_ANALYSIS_STDERR,
    COL_ANALYSIS_UPDATED,
    COL_CONVERSION_STATUS,
    COL_CONVERSION_RETURN,
    COL_CONVERSION_STDOUT,
    COL_CONVERSION_STDERR,
    COL_CONVERSION_UPDATED,
    COL_RETRY_COUNT,
]


def yes_no(value: bool | None) -> str:
    if value is None:
        return ""
    return "是" if value else "否"


def active_manual_comment(value: str) -> bool:
    text = (value or "").strip()
    return bool(text) and not text.startswith("已完成")


def ensure_headers(headers: list[str]) -> list[str]:
    out = []
    seen = set()
    for header in headers + DEFAULT_COLUMNS:
        header = str(header).strip()
        if not header or header in seen:
            continue
        seen.add(header)
        out.append(header)
    return out


@contextlib.contextmanager
def board_lock(path: pathlib.Path | str) -> Iterator[None]:
    board_path = pathlib.Path(path)
    board_path.parent.mkdir(parents=True, exist_ok=True)
    lock_path = board_path.with_suffix(board_path.suffix + ".lock")
    fd = os.open(lock_path, os.O_CREAT | os.O_RDWR, 0o666)
    try:
        import fcntl

        fcntl.flock(fd, fcntl.LOCK_EX)
        yield
    finally:
        import fcntl

        fcntl.flock(fd, fcntl.LOCK_UN)
        os.close(fd)


def _open_workbook(path: pathlib.Path) -> tuple[Workbook, Worksheet]:
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = SHEET_NAME
        return workbook, sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    return workbook, sheet


def _headers_from_sheet(sheet: Worksheet) -> list[str]:
    headers = []
    for cell in sheet[1]:
        value = str(cell.value or "").strip()
        if value:
            headers.append(value)
    return ensure_headers(headers)


def _rows_from_sheet(sheet: Worksheet, headers: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if sheet.max_row < 2:
        return rows
    for values in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers), values_only=True):
        row = {header: str(values[index] or "") for index, header in enumerate(headers)}
        if any(value.strip() for value in row.values()):
            rows.append(row)
    return rows


def load_board(path: pathlib.Path | str) -> tuple[list[str], list[dict[str, str]]]:
    board_path = pathlib.Path(path)
    if not board_path.exists():
        return DEFAULT_COLUMNS[:], []
    workbook, sheet = _open_workbook(board_path)
    headers = _headers_from_sheet(sheet)
    rows = _rows_from_sheet(sheet, headers)
    workbook.close()
    return headers, rows


def _apply_sheet_style(sheet: Worksheet, headers: list[str], row_count: int) -> None:
    sheet.freeze_panes = "A2"
    if headers:
        sheet.auto_filter.ref = f"A1:{sheet.cell(row=max(1, row_count + 1), column=len(headers)).coordinate}"
    wide = {COL_MK_PATH, COL_CMAKE_PATH, COL_MANUAL_COMMENT, COL_ANALYSIS_RESULT, COL_ANALYSIS_STDOUT, COL_ANALYSIS_STDERR, COL_CONVERSION_STDOUT, COL_CONVERSION_STDERR}
    for index, header in enumerate(headers, start=1):
        column = sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter]
        column.width = 48 if header in wide else 18
    for cell in sheet[1]:
        cell.style = "Headline 4"


def save_board(path: pathlib.Path | str, headers: list[str], rows: list[dict[str, str]]) -> None:
    board_path = pathlib.Path(path)
    board_path.parent.mkdir(parents=True, exist_ok=True)
    headers = ensure_headers(headers)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(headers)
    for row in rows:
        sheet.append([str(row.get(header, "") or "") for header in headers])
    _apply_sheet_style(sheet, headers, len(rows))

    with tempfile.NamedTemporaryFile(delete=False, dir=str(board_path.parent), suffix=".xlsx.tmp") as tmp:
        tmp_path = pathlib.Path(tmp.name)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, board_path)
    finally:
        workbook.close()
        if tmp_path.exists():
            tmp_path.unlink()


def mutate_board(
    path: pathlib.Path | str,
    mutator: Callable[[list[str], list[dict[str, str]]], tuple[list[str], list[dict[str, str]]] | None],
) -> tuple[list[str], list[dict[str, str]]]:
    with board_lock(path):
        headers, rows = load_board(path)
        result = mutator(headers, rows)
        if result is not None:
            headers, rows = result
        save_board(path, headers, rows)
        return headers, rows


def row_by_path(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {row.get(COL_MK_PATH, ""): row for row in rows if row.get(COL_MK_PATH, "")}


def row_by_id(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {row.get(COL_TASK_ID, ""): row for row in rows if row.get(COL_TASK_ID, "")}
